from openerp import models, fields, api, _
from openerp.exceptions import ValidationError
from datetime import datetime, date
from openerp.tools.misc import DEFAULT_SERVER_DATE_FORMAT
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
import os
from cStringIO import StringIO
import logging
_logger = logging.getLogger(__name__)

class BankLedgerWizard(models.TransientModel):
    _name = "bank.ledger.wizard"

    @api.model
    def _default_society(self):
        user_id = self.env['res.users'].sudo().browse(self.env.uid)
        if len(user_id.company_id.parent_id) > 0 and user_id.company_id.parent_id.type == 'society':
            return user_id.company_id.parent_id.id
        elif user_id.company_id.type == 'society':
            return user_id.company_id.id

    @api.model
    def _default_school(self):
        user_id = self.env['res.users'].sudo().browse(self.env.uid)
        if user_id.company_id and user_id.company_id.type == 'school':
            return user_id.company_id.id

    society_id =fields.Many2one('res.company',string= 'Society', default=_default_society)
    school_id = fields.Many2one('res.company',string='School', default=_default_school)
    academic_year_id = fields.Many2one('academic.year',string='Academic Year')
    bank_account_id = fields.Many2one('bank.account.config', string='Bank Name')
    from_date = fields.Date(string='From Date')
    to_date = fields.Date(string='To Date')

#     @api.constrains('from_date','to_date')
#     def check_date(self):
#         if self.from_date and self.academic_year_id and self.from_date < self.academic_year_id.start_date:
#             raise ValidationError(_('From Date should be greater than Academic Year start date!'))
#         elif self.to_date and self.academic_year_id and self.to_date > self.academic_year_id.end_date:
#             raise ValidationError(_('To Date should not be greater than Academic Year End date!'))
#         elif self.from_date and self.to_date and self.from_date > self.to_date:
#             raise ValidationError(_('To Date should be greater than the From Date!'))

    @api.onchange('society_id')
    def onchange_society(self):
        if self.society_id:
            self.school_id = None
            self.academic_year_id = None
            self.bank_account_id = None
            self.from_date = None
            self.to_date = None

    @api.onchange('school_id')
    def onchange_school(self):
        if self.school_id:
            self.academic_year_id = None
            self.bank_account_id = None
            self.from_date = None
            self.to_date = None

    @api.onchange('academic_year_id')
    def onchange_academic_year(self):
        if self.academic_year_id:
            self.bank_account_id = None
            self.from_date = None
            self.to_date = None

    @api.onchange('from_date')
    def onchange_date(self):
        if self.from_date:
            self.to_date = None

    @api.multi
    def generate_xl(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "CASH BOOK"
        ws.append([(self.school_id.name if self.school_id.name else '')])
        ws.append([(self.school_id.street if self.school_id.street else '') + ', ' + (self.school_id.street2 if self.school_id.street2 else '') + ', ' + (self.school_id.city if self.school_id.city else '')])
        ws.append(['P.O Box: ' + (self.school_id.street2 if self.school_id.street2 else '')])
        ws.append(['Tel: ' + (self.school_id.mobile if self.school_id.mobile else '') + ', ' + 'Fax: ' + (self.school_id.fax_id if self.school_id.fax_id else '') + ', ' + 'Email: ' + (self.school_id.email if self.school_id.email else '')])
        ws.append([self.school_id.website if self.school_id.website else ''])
        ws.append([])
        ws.append(['CASH BOOK'])
        ws.append(['' + (datetime.strptime(str(self.from_date),DEFAULT_SERVER_DATE_FORMAT).strftime('%d/%b/%Y') + ' - ' + datetime.strptime(str(self.to_date),DEFAULT_SERVER_DATE_FORMAT).strftime('%d/%b/%Y'))])
        ws.append([])
        ws.append(['Date', 'Opening Balance', 'Cash Collection', 'Bank Deposit', 'Closing Balance', 'Uncleared Deposit'])
        t_count = 10
        deposit_obj = self.env['bank.deposit'].sudo().search([('society_id','=',self.society_id.id),('school_id','=',self.school_id.id),
                                                       ('academic_year_id', '=', self.academic_year_id.id),('bank_id', '=', self.bank_account_id.id),
                                                       ('deposit_date', '>=', self.from_date),('deposit_date', '<=', self.to_date),('state','in',['requested','approved'])])
        for obj in deposit_obj:
            unclr,clr = 0.0,0.0
            date = datetime.strptime(str(obj.deposit_date),DEFAULT_SERVER_DATE_FORMAT).strftime('%d-%b-%Y')
            if obj.state == 'requested':
                unclr += obj.c_amt_deposit
            if obj.state == 'approved':
                clr += obj.c_amt_deposit
            ws.append([date,obj.opening_bal,obj.total_cash_amt,clr,obj.closing_bal,unclr])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['C' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['D' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['E' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
            ws.cell(row=t_count + 1, column=1).border = thin_border
            ws.cell(row=t_count + 1, column=2).border = thin_border
            ws.cell(row=t_count + 1, column=3).border = thin_border
            ws.cell(row=t_count + 1, column=4).border = thin_border
            ws.cell(row=t_count + 1, column=5).border = thin_border
            ws.cell(row=t_count + 1, column=6).border = thin_border
            t_count += 1

        # Company Details
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ft2 = Font(size=10)
        header2 = NamedStyle(name="header2", font=ft2)
        ft3 = Font(size=11, bold=True)
        header3 = NamedStyle(name="header3", font=ft3)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:F2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:F3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:F4')
        ws['A5'].style = header2
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:F5')
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:F6')
        # styles for row7
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:F7')
        # Styles for row 8,9
        ws['A8'].style = header2
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        # Styles for row 10
        ws['A10'].style = header3
        ws['A10'].alignment = Alignment(horizontal="center")
        ws['B10'].style = header3
        ws['B10'].alignment = Alignment(horizontal="center")
        ws['C10'].style = header3
        ws['C10'].alignment = Alignment(horizontal="center")
        ws['D10'].style = header3
        ws['D10'].alignment = Alignment(horizontal="center")
        ws['E10'].style = header3
        ws['E10'].alignment = Alignment(horizontal="center")
        ws['F10'].style = header3
        ws['F10'].alignment = Alignment(horizontal="center")
        # Borders
        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
        ws.cell(row=10, column=1).border = thin_border
        ws.cell(row=10, column=2).border = thin_border
        ws.cell(row=10, column=3).border = thin_border
        ws.cell(row=10, column=4).border = thin_border
        ws.cell(row=10, column=5).border = thin_border
        ws.cell(row=10, column=6).border = thin_border
        # Width style
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        fp = StringIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_excel_report(self):
        # if not self.get_bank_ledger():
        #     raise ValidationError(_("No record found..!"))
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Cash Book'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Cash Book'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }